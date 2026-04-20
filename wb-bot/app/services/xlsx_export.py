from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


HEADER = [
    "Артикул Продавца",
    "Заказы, шт",
    "Заказы, сумма",
    "Остатки, шт",
    "Динамика заказов в рублях",
]


def autosize(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            width = len(str(value or "")) + 2
            widths[idx] = max(widths.get(idx, 0), min(width, 50))
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def save_sales_report(rows: Iterable[dict], output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Продажи"
    ws.append(HEADER)

    for row in rows:
        ws.append(
            [
                row.get("vendor_article", ""),
                row.get("orders_qty", 0),
                row.get("orders_sum", 0),
                row.get("stock_qty", 0),
                row.get("orders_sum_dynamic", 0),
            ]
        )

    autosize(ws)
    wb.save(output_path)
    return output_path
