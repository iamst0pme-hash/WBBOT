from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def to_decimal(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal("0")


@dataclass
class SalesRow:
    article: str
    orders_qty: Decimal
    orders_amount: Decimal
    stock_qty: Decimal
    dynamics_amount: Decimal


def build_sales_workbook(rows: Iterable[SalesRow], period_label: str) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Продажи"

    ws["A1"] = "Временная выгрузка по продажам"
    ws["A2"] = f"Период: {period_label}"

    headers = [
        "Артикул Продавца",
        "Заказы, шт",
        "Заказы, сумма",
        "Остатки, шт",
        "Динамика заказов в рублях",
    ]
    header_row = 4
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, start=header_row + 1):
        ws.cell(row=row_idx, column=1, value=row.article)
        ws.cell(row=row_idx, column=2, value=float(row.orders_qty))
        ws.cell(row=row_idx, column=3, value=float(row.orders_amount))
        ws.cell(row=row_idx, column=4, value=float(row.stock_qty))
        ws.cell(row=row_idx, column=5, value=float(row.dynamics_amount))

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:E{max(header_row + 1, ws.max_row)}"

    for col in ["B", "C", "D", "E"]:
        for cell in ws[col][header_row:]:
            cell.number_format = '#,##0.00'

    widths = {
        "A": 28,
        "B": 16,
        "C": 18,
        "D": 16,
        "E": 24,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.column_dimensions["A"].width = 28
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"].font = Font(italic=True)

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        return Path(tmp.name)
